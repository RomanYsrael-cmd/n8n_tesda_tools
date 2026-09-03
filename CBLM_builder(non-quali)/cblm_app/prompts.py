from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


PROMPT_ROWS = {
    "module_title": 5,
    "module_descriptor": 11,
    "lesson_objectives": 20,
    "keyfacts_content": 21,
    "self_check": 23,
    "task_sheet": 27,
}


@dataclass(frozen=True)
class PromptCatalog:
    prompts: dict[str, str]

    @classmethod
    def load(cls, path: Path) -> "PromptCatalog":
        sheet = load_workbook(path, read_only=True, data_only=False)["Sheet1"]
        prompts = {name: str(sheet.cell(row=row, column=2).value or "") for name, row in PROMPT_ROWS.items()}
        missing = [name for name, prompt in prompts.items() if not prompt.strip()]
        if missing:
            raise ValueError("Prompts.xlsx has empty prompts: " + ", ".join(missing))
        return cls(prompts)

    def render(self, name: str, values: dict[str, object]) -> str:
        prompt = self.prompts[name]
        lookup = {key.casefold(): _full_value(value) for key, value in values.items()}

        def replace(match: re.Match) -> str:
            key = match.group(1).strip().casefold()
            return lookup.get(key, match.group(0))

        return re.sub(r"\{\{([^{}]+)\}\}", replace, prompt)


def _full_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "\n".join(str(item) for item in value)
    return str(value)

